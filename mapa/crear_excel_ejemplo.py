"""
Script para crear archivo Excel de ejemplo para MapGlow
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Crear nuevo workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ubicaciones"

# Estilos para encabezados - colores morados/rosas
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='9333ea', end_color='9333ea', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

# Configurar encabezados
ws['A1'] = 'Descripción'
ws['B1'] = 'Coordenadas'

ws['A1'].font = header_font
ws['A1'].fill = header_fill
ws['A1'].alignment = header_alignment

ws['B1'].font = header_font
ws['B1'].fill = header_fill
ws['B1'].alignment = header_alignment

# Ajustar anchos de columnas
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 22

# Datos de ejemplo - Lugares bonitos y variados
ubicaciones = [
    ("Plaza de Armas", "-6.6408, -79.3875"),
    ("Parque Principal", "-6.6425, -79.3890"),
    ("Centro Comercial", "-6.6395, -79.3862"),
    ("Cafetería del Centro", "-6.6418, -79.3868"),
    ("Instituto Tecnológico", "-6.6438, -79.3905"),
    ("Biblioteca Municipal", "-6.6410, -79.3873"),
]

# Agregar datos
for i, (descripcion, coordenadas) in enumerate(ubicaciones, start=2):
    ws[f'A{i}'] = descripcion
    ws[f'B{i}'] = coordenadas
    ws[f'A{i}'].alignment = Alignment(vertical='center')
    ws[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')

# Guardar archivo
filename = 'ubicaciones_ejemplo.xlsx'
wb.save(filename)

print(f"✨ Archivo '{filename}' creado exitosamente")
print(f"\n💝 Se incluyeron {len(ubicaciones)} ubicaciones:")
for i, (desc, coords) in enumerate(ubicaciones, 1):
    print(f"   {i}. {desc} → {coords}")
print("\n🌸 ¡Listo para usar en MapGlow!")
